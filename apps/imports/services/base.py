"""
Base Importer - Clase base para todos los importadores.
"""
import os
import csv
import io
from abc import ABC, abstractmethod
from django.db import transaction
from django.utils import timezone
import openpyxl

from ..models import ImportLog, ImportError


class BaseImporter(ABC):
    """Clase base abstracta para importadores."""
    
    TIPO = None  # 'productos', 'clientes', 'categorias'
    COLUMNAS_REQUERIDAS = []  # Lista de columnas obligatorias
    COLUMNAS_OPCIONALES = []  # Lista de columnas opcionales
    BATCH_SIZE = 500  # Tamaño de lote para procesamiento
    
    def __init__(self, archivo, usuario=None, opciones=None):
        """
        Inicializa el importador.
        
        Args:
            archivo: Archivo subido (InMemoryUploadedFile o path)
            usuario: Usuario que realiza la importación
            opciones: Dict con opciones adicionales
        """
        self.archivo = archivo
        self.usuario = usuario
        self.opciones = opciones or {}
        self.datos = []
        self.errores = []
        self.preview_data = {
            'filas': [],
            'a_crear': 0,
            'a_actualizar': 0,
            'errores': [],
            'total': 0
        }
        self.log = None
    
    def leer_archivo(self):
        """Lee el archivo Excel o CSV y retorna lista de diccionarios."""
        nombre = self.archivo.name if hasattr(self.archivo, 'name') else str(self.archivo)
        extension = os.path.splitext(nombre)[1].lower()
        
        if extension in ['.xlsx', '.xls']:
            return self._leer_excel()
        elif extension == '.csv':
            return self._leer_csv()
        else:
            raise ValueError(f"Formato no soportado: {extension}. Use .xlsx o .csv")
    
    def _leer_excel(self):
        """Lee archivo Excel usando openpyxl."""
        if hasattr(self.archivo, 'read'):
            wb = openpyxl.load_workbook(self.archivo, data_only=True)
        else:
            wb = openpyxl.load_workbook(self.archivo, data_only=True)
        
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        
        if not rows:
            return []
        
        # Primera fila son los headers
        headers = [str(h).strip() if h else '' for h in rows[0]]
        
        datos = []
        for i, row in enumerate(rows[1:], start=2):
            if all(cell is None or str(cell).strip() == '' for cell in row):
                continue  # Saltar filas vacías
            
            fila_dict = {'_fila': i}
            for j, header in enumerate(headers):
                if header and j < len(row):
                    valor = row[j]
                    if valor is not None:
                        fila_dict[header] = str(valor).strip() if not isinstance(valor, (int, float)) else valor
                    else:
                        fila_dict[header] = ''
            datos.append(fila_dict)
        
        return datos
    
    def _leer_csv(self):
        """Lee archivo CSV."""
        if hasattr(self.archivo, 'read'):
            content = self.archivo.read()
            if isinstance(content, bytes):
                content = content.decode('utf-8-sig')
            reader = csv.DictReader(io.StringIO(content))
        else:
            with open(self.archivo, 'r', encoding='utf-8-sig') as f:
                reader = csv.DictReader(f)
                return [{'_fila': i+2, **row} for i, row in enumerate(reader)]
        
        datos = []
        for i, row in enumerate(reader, start=2):
            row['_fila'] = i
            datos.append(row)
        return datos
    
    def validar_columnas(self, datos):
        """Valida que existan las columnas requeridas (case-insensitive)."""
        if not datos:
            raise ValueError("El archivo está vacío")
        
        # Obtener columnas del archivo en lowercase para comparación
        columnas_archivo = {k.lower().strip() for k in datos[0].keys() if k != '_fila'}
        
        # Verificar columnas requeridas (case-insensitive)
        faltantes = []
        for col_req in self.COLUMNAS_REQUERIDAS:
            if col_req.lower().strip() not in columnas_archivo:
                faltantes.append(col_req)
        
        if faltantes:
            raise ValueError(f"Columnas faltantes: {', '.join(faltantes)}")
    
    def preview(self):
        """
        Realiza un dry-run y retorna el preview.
        
        Returns:
            dict con: filas (primeras 10), a_crear, a_actualizar, errores, total
        """
        self.datos = self.leer_archivo()
        self.validar_columnas(self.datos)
        
        self.preview_data = {
            'filas': self.datos[:10],
            'a_crear': 0,
            'a_actualizar': 0,
            'errores': [],
            'total': len(self.datos)
        }
        
        for fila in self.datos:
            try:
                accion, obj = self.procesar_fila(fila, dry_run=True)
                if accion == 'crear':
                    self.preview_data['a_crear'] += 1
                elif accion == 'actualizar':
                    self.preview_data['a_actualizar'] += 1
            except Exception as e:
                self.preview_data['errores'].append({
                    'fila': fila.get('_fila', '?'),
                    'mensaje': str(e)
                })
        
        return self.preview_data
    
    def ejecutar(self):
        """
        Ejecuta la importación real.
        
        Returns:
            ImportLog con el resultado
        """
        # Crear log
        self.log = ImportLog.objects.create(
            tipo=self.TIPO,
            estado='procesando',
            archivo_nombre=self.archivo.name if hasattr(self.archivo, 'name') else str(self.archivo),
            usuario=self.usuario,
            total_filas=len(self.datos)
        )
        
        try:
            creados = 0
            actualizados = 0
            errores = 0
            
            for i, fila in enumerate(self.datos):
                try:
                    accion, obj = self.procesar_fila(fila, dry_run=False)
                    if accion == 'crear':
                        creados += 1
                    elif accion == 'actualizar':
                        actualizados += 1
                except Exception as e:
                    errores += 1
                    ImportError.objects.create(
                        log=self.log,
                        fila=fila.get('_fila', i+2),
                        mensaje=str(e)
                    )
                
                # Actualizar progreso cada 5 filas para feedback más fluido
                if (i + 1) % 5 == 0:
                    self.log.procesados = i + 1
                    self.log.save(update_fields=['procesados'])
            
            # Finalizar
            self.log.creados = creados
            self.log.actualizados = actualizados
            self.log.errores = errores
            self.log.procesados = len(self.datos)
            self.log.estado = 'completado'
            self.log.completed_at = timezone.now()
            self.log.save()
            
        except Exception as e:
            self.log.estado = 'error'
            self.log.save()
            raise
        
        return self.log
    
    @abstractmethod
    def procesar_fila(self, fila, dry_run=False):
        """
        Procesa una fila individual.
        
        Args:
            fila: Dict con los datos de la fila
            dry_run: Si es True, no guarda cambios
        
        Returns:
            Tuple (accion, objeto) donde accion es 'crear', 'actualizar' o 'saltar'
        """
        pass
    
    def get_valor(self, fila, columna, default=''):
        """Obtiene un valor de la fila, manejando variaciones en nombres."""
        # Buscar exacto
        if columna in fila:
            return fila[columna]
        
        # Buscar ignorando mayúsculas/espacios
        columna_lower = columna.lower().strip()
        for key in fila:
            if key.lower().strip() == columna_lower:
                return fila[key]
        
        return default
    
    def get_decimal(self, fila, columna, default=0):
        """Obtiene un valor decimal de la fila."""
        valor = self.get_valor(fila, columna, '')
        if valor == '' or valor is None:
            return default
        
        try:
            # Manejar formatos como "10%", "10,5", etc.
            valor_str = str(valor).replace('%', '').replace(',', '.').strip()
            return float(valor_str)
        except (ValueError, TypeError):
            raise ValueError(f"'{valor}' no es un número válido en columna {columna}")
    
    def get_int(self, fila, columna, default=0):
        """Obtiene un valor entero de la fila."""
        valor = self.get_valor(fila, columna, '')
        if valor == '' or valor is None:
            return default
        
        try:
            return int(float(str(valor).replace(',', '.').strip()))
        except (ValueError, TypeError):
            raise ValueError(f"'{valor}' no es un número entero válido en columna {columna}")
    
    def generar_csv_errores(self):
        """Genera un CSV con los errores de la importación."""
        if not self.log:
            return None
        
        output = io.StringIO()
        writer = csv.writer(output)
        writer.writerow(['Fila', 'Columna', 'Valor', 'Error'])
        
        for error in self.log.errores_detalle.all():
            writer.writerow([error.fila, error.columna, error.valor, error.mensaje])
        
        return output.getvalue()
